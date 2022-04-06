import openpyxl
import time
import datetime
import PySimpleGUI as sg
import pickle

def read_info(cw, day, add = 'Work.xlsx'):
    # ブック、シートを開く
    book = openpyxl.load_workbook(add)
    sheet = book.active
    info = sheet.cell(day+1, cw+1).value
    return info

def write_info(cw, day, info, add = 'Work.xlsx'):
    book = openpyxl.load_workbook(add)
    sheet = book.active
    sheet.cell(day+1, cw+1).value = info
    book.save(add)
    return

def save_database(data, *, pkl_file = 'database.pkl'):
    with open(pkl_file, 'wb') as f:
        pickle.dump(data, f)
    return

def read_database(*, read_all = True, cw = None, day = None, pkl_file = 'database.pkl'):
    with open(pkl_file, 'rb') as f:
        data = pickle.load(f)
    if read_all == False:
        return data[day- 1][cw - 1]
    else:
        return data

def isfloat(txt):
    try:
        float(txt)  
        return True
    except ValueError:
        return False
    
class UserInterfaces:
    def __init__(self,database):
        self.db = database
        sg.theme('SystemDefault')
        self.window_startgui = None
        self.window_update_task = None

    def start_gui(self):
        button1 = sg.Button('Today', key = '-today-', size = (5,1), font=('Arial',12))
        
        week_days = ['Monday', ' Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        combo = sg.Combo(week_days, default_value = 'Choose day of week', size = (20,1), readonly = True, key = '-day-', enable_events = True, font=('Arial',16))

        text = sg.Text('Calendar week: ', font=('Arial',16))
        
        input = sg.InputText(size = (10,1), key='-cw-', font=('Arial',16))

        button2 = sg.Button('Next', key = '-next-', size = (5,1), font=('Arial',12))

        layout = [[text, input, combo], [button1, button2]]
        self.window_startgui = sg.Window('Task Manager --Welcome', layout)

        today_date = datetime.date.today()
        today_week = today_date.isocalendar()[1]
        today_day = today_date.isocalendar()[2]

        while True:  # Event Loop
            event, values = self.window_startgui.read()
            if event == sg.WIN_CLOSED:
                break
            if event == '-today-':
                self.window_startgui['-day-'].Update(value = week_days[today_day-1])
                self.window_startgui['-cw-'].Update(value = today_week)
            if event == '-next-':
                if (values['-day-'] in week_days) and (values['-cw-'].isdigit()) and (1 <= int(values['-cw-']) <= 52):
                    #self.window_startgui.close()
                    #self.task_info_gui(int(values['-cw-']), int(week_days.index(values['-day-']) + 1))
                    self.input_task_gui(int(values['-cw-']), int(week_days.index(values['-day-']) + 1))
                    #break
                else:
                    sg.popup('Wrong Input!', font=('Arial',12))
                    continue
        self.window_startgui.close()
        

    def task_info_gui(self, cw_this = None, day_this = None):
        daily_info = self.db[day_this - 1][cw_this - 1]
        #print(daily_info.generate_txt())
        return

    def input_task_gui(self, cw_this = None, day_this = None):
        daily_info = self.db[day_this - 1][cw_this - 1]
        if daily_info != None:
            task_num = len(daily_info.total_task)
        else:
            task_num = 0
        
        total_task = 8

        empty_text = sg.Text('', font=('Arial',12), size = (3,1))
        text_name = sg.Text('Task Name: ', font=('Arial',12), size = (25,1))
        text_plan = sg.Text('Plan/(h): ', font=('Arial',12), size = (8,1))
        text1_actual = sg.Text('Actual/(h): ', font=('Arial',12), size = (8,1))
        text_comment = sg.Text('Comment: ', font=('Arial',12), size = (20,1))
        row0 = [empty_text, text_name, text_plan, text1_actual, text_comment]
        
        for index in range(1,total_task + 1):
            if index <= task_num:
                globals()[f'input{index}_0'] = sg.InputText(default_text = str(daily_info.total_task[index-1].task_id), key=f'-id{index}-', font=('Arial',12), size = (3,1))
                globals()[f'input{index}_1'] = sg.InputText(default_text = str(daily_info.total_task[index-1].task_name), key=f'-name{index}-', font=('Arial',12), size = (25,1))
                globals()[f'input{index}_2'] = sg.InputText(default_text = str(daily_info.total_task[index-1].task_plan_time), key=f'-plantime{index}-', font=('Arial',12), size = (8,1))
                globals()[f'input{index}_3'] = sg.InputText(default_text = str(daily_info.total_task[index-1].task_actual_time), key=f'-actualtime{index}-', font=('Arial',12), size = (8,1))
                globals()[f'input{index}_4'] = sg.InputText(default_text = str(daily_info.total_task[index-1].task_comment), key=f'-comment{index}-', font=('Arial',12), size = (20,1))
                globals()[f'row{index}'] = [globals()[f'input{index}_0'], globals()[f'input{index}_1'], globals()[f'input{index}_2'], globals()[f'input{index}_3'], globals()[f'input{index}_4']]
            else:
                globals()[f'input{index}_0'] = sg.InputText('', key=f'-id{index}-', font=('Arial',12), size = (3,1))
                globals()[f'input{index}_1'] = sg.InputText('',key=f'-name{index}-', font=('Arial',12), size = (25,1))
                globals()[f'input{index}_2'] = sg.InputText('', key=f'-plantime{index}-', font=('Arial',12), size = (8,1))
                globals()[f'input{index}_3'] = sg.InputText('', key=f'-actualtime{index}-', font=('Arial',12), size = (8,1))
                globals()[f'input{index}_4'] = sg.InputText('', key=f'-comment{index}-', font=('Arial',12), size = (20,1))
                globals()[f'row{index}'] = [globals()[f'input{index}_0'], globals()[f'input{index}_1'], globals()[f'input{index}_2'], globals()[f'input{index}_3'], globals()[f'input{index}_4']]
        
        button1 = sg.Button('Update', key = '-update-', size = (6,1), font=('Arial',12))
        button2 = sg.Button('Clear', key = '-clear-', size = (6,1), font=('Arial',12))
        #button3 = sg.Button('Back', key = '-back-', size = (6,1), font=('Arial',12))

        layout = [row0, row1, row2, row3, row4, row5, row6, row7, row8],[button1,button2] # type: ignore 
        
        self.window_update_task = sg.Window('Task Manager --Update Task', layout)
        
        while True:  # Event Loop
            event, values = self.window_update_task.read()
            if event == sg.WIN_CLOSED:
                break

            if event == '-update-':
                if daily_info != None:
                    daily_info.clear_task()
                else:
                    daily_info = DailyInfo(cw = cw_this, day = day_this)
            
                for index in range(1,total_task + 1):
                    new_name = values[f'-name{index}-']
                    new_id = values[f'-id{index}-']
                    new_plan_time = values[f'-plantime{index}-']
                    new_actual_time = values[f'-actualtime{index}-']
                    new_comment = values[f'-comment{index}-']
                    if (new_name != '') or (new_id != '') and (new_plan_time != '') or (new_actual_time != '') or (new_comment != ''):
                        if new_id == '':
                            new_id = '0'
                        if new_plan_time == '':
                            new_plan_time = '0.0'
                        if new_actual_time == '':
                            new_actual_time = '0.0'
                        if not isfloat(new_plan_time):
                            sg.popup(f'Wrong Plan Time Input! ({new_plan_time})', font=('Arial',12))
                            new_plan_time = '0.0'
                        if not isfloat(new_actual_time):
                            sg.popup(f'Wrong Actual Time Input! ({new_actual_time})', font=('Arial',12))
                            new_actual_time = '0.0'
                            
                        globals()[f'task{index}'] = Task(name = new_name, id = int(new_id), plan_time = float(new_plan_time), actual_time = float(new_actual_time), comment = new_comment)
                        daily_info.add_task(globals()[f'task{index}'])

                self.db[day_this - 1][cw_this - 1] = daily_info
                save_database(self.db)

            if event == '-clear-':
                for index in range(1,total_task + 1):
                    self.window_update_task[f'-name{index}-'].Update(value = '')
                    self.window_update_task[f'-id{index}-'].Update(value = '')
                    self.window_update_task[f'-plantime{index}-'].Update(value = '')
                    self.window_update_task[f'-actualtime{index}-'].Update(value = '')
                    self.window_update_task[f'-comment{index}-'].Update(value = '')
            '''
            if event == '-back-':
                self.window_update_task.close()
                self.start_gui()
                break
            '''     
        self.window_update_task.close()   
        
        #print(daily_info.generate_txt())
        return

class Task:
    def __init__(self, *, name = '', plan_time = 0.0, actual_time = 0.0, comment = '', id = 0):
        self.task_name = name
        self.task_plan_time = plan_time
        self.task_actual_time = actual_time
        self.task_comment = comment
        self.start_time = None
        self.elapsed_time = None
        self.task_id = id

    def read(self):
        return (self.task_name, self.task_plan_time, self.task_actual_time, self.task_comment)

    def write(self, name = None, plan_time = None, actual_time = None, comment = None):
        if name is not None:
            self.task_name = name
        if plan_time is not None:
            self.task_plan_time = plan_time
        if actual_time is not None:
            self.task_actual_time = actual_time
        if comment is not None:
            self.task_comment = comment
        return

    def count_actual_time(self):
        self.start_time = time.time()
        #print('count start')
        return
    
    def update_actual_time(self):
        if self.start_time is not None:
            self.elapsed_time = int(time.time() - self.start_time)
            elapsed_hour = round(self.elapsed_time/3600, 4)
            self.task_actual_time += elapsed_hour
            #print('count time = ' + str(self.task_actual_time) + 'hour')
        return

class DailyInfo:
    def __init__(self, cw = 0, day = 0):
        self.cw = cw
        self.day = day
        self.total_task = []
        t_delta = datetime.timedelta(hours=9)
        JST = datetime.timezone(t_delta, 'JST')
        now = datetime.datetime.now(JST)
        datestr = now.strftime('%Y/%m/%d')
        self.txt = datestr + '\n'

    def add_task(self,task):
        self.total_task.append(task)
        return

    def clear_task(self):
        self.total_task = []
        return

    def generate_txt(self):       
        for task_in_list in self.total_task:
            task_txt = str(task_in_list.task_id) + '. ' + task_in_list.task_name + ' (' + str(task_in_list.task_plan_time) + 'h)' + '(✔️' + str(task_in_list.task_actual_time) + 'h)' + ' ' + task_in_list.task_comment + '\n'
            self.txt += task_txt
        return self.txt


def main():
    '''
    task1 = Task(name = 'MTG', comment = 'daily', plan_time = 4.0, id = 1)
    task1.write(actual_time = 2.0)
    task2 = Task(name = 'python programming', plan_time = 2.0, actual_time = 2.0, id = 2)
    task3 = Task(name = 'pc setup', plan_time = 1.0, actual_time = 1.5, id = 3)

    day1 = DailyInfo(day = 1, cw = 1)
    day1.add_task(task1)
    day1.add_task(task2)
    day1.add_task(task3)
    #write_info(cw = day1.cw, day = day1.day, info = day1.generate_txt())

    row, column = (60, 7) 
    database = [[None for i in range(row)] for j in range(column)]
    cw_db = 1
    day_db = 1
    database[cw_db - 1][day_db - 1] = day1
    '''
    
    database = read_database()

    ui = UserInterfaces(database)
    ui.start_gui()


    save_database(database)  

if __name__ == "__main__":
    main()