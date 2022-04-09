import openpyxl
import pickle
import datetime

# read information from an excel cell defined by calendar week and day of week
def read_info(cw, day, add = 'Work.xlsx'):
    book = openpyxl.load_workbook(add)
    sheet = book.active
    info = sheet.cell(day+1, cw+1).value
    return info

# read information to an excel cell defined by calendar week and day of week
def write_info(cw, day, info, add = 'Work.xlsx'):
    book = openpyxl.load_workbook(add)
    sheet = book.active
    sheet.cell(day+1, cw+1).value = info
    book.save(add)

# save database to a .pkl file
# database is a 2d list with DailyInfo class data in each element
def save_database(data, *, pkl_file = 'database.pkl'):
    with open(pkl_file, 'wb') as f:
        pickle.dump(data, f)

# load database from a .pkl file
# database is a 2d list with DailyInfo class data in each element
# when "read_all = True", output is the whole 2d list
# when "read_all = False", output is the DailyInfo class data in the 2d list
def read_database(*, read_all = True, cw = None, day = None, pkl_file = 'database.pkl'):
    with open(pkl_file, 'rb') as f:
        data = pickle.load(f)
    if read_all == False:
        return data[day- 1][cw - 1]
    else:
        return data

# a function to check if the input can be converted to float type data
# "True" doesn't mean that input IS float
def isfloat(txt):
    try:
        float(txt)  
        return True
    except ValueError:
        return False

# a function to get calendar week and day of week of today 
def get_today():
        date = datetime.date.today()
        week = date.isocalendar()[1]
        day = date.isocalendar()[2]
        return week, day
